# Import the load_workbook function from openpyxl to work with Excel files
from openpyxl import load_workbook
# Import nltk for natural language processing tasks
import nltk

# Specify the path to your Excel file
file_path = '../dataset/dataset.xlsx'

# Load the workbook
wb = load_workbook(filename=file_path)

# Access a specific sheet named 'Dataset'
sheet = wb['Dataset']

# Close the workbook after loading the required sheet
wb.close()

# Download necessary NLTK data
nltk.download('punkt')         # Tokenizer for splitting sentences into words
nltk.download('stopwords')     # Common stop words like 'the', 'and', 'is', etc.
nltk.download('wordnet')       # Lexical database for English language (for lemmatization)
nltk.download('words')         # List of words in the English language


